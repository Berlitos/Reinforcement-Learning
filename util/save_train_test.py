import matplotlib.pyplot as plt
import os
from openpyxl import Workbook, load_workbook
import sys

def save_plot_reward(algo, env, history, model, params):
    ## Save plot image ##
    plt.plot(history.history['episode_reward'])
    plt.title('reward training model ' + model)
    plt.ylabel('reward')
    plt.xlabel('episode')
    plt.savefig('../' + algo + '/' + env + '/plots/' + model + '.png')
    
    ## Save params excel ##
    filepath = '../' + algo + '/' + env + '/results.xlsx'
    wb = load_workbook(filepath)
    sheet = wb.active
    sheet.cell(row=int(model)+1, column=1).value = int(model)
    for y, param in enumerate(params):
        sheet.cell(row=int(model)+1, column=y+2).value = param
    wb.save(filepath)
    plt.show()
    wb.close()

def save_result(algo, env, history, model, params):
    rewards = history.history['episode_reward']
    result = sum(rewards) / float(len(rewards))
    filepath = '../' + algo + '/' + env + '/results.xlsx'
    wb = load_workbook(filepath)
    sheet = wb.active
    sheet.cell(row=int(model)+1, column=len(params) + 2).value = result
    wb.save(filepath)
    wb.close()