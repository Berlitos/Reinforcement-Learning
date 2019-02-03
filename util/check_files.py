import matplotlib.pyplot as plt
import os
from openpyxl import Workbook, load_workbook
import sys


def check(algo, env):
    check_env(algo, env)
    check_xml(algo, env)


def check_env(algo, env):
    choice = 'null'
    filepath = '../' + algo + '/' + env
    if not os.path.isdir(filepath):
        while(choice != 'y' and choice != 'n'):
            choice = input('Il n\'existe pas encore de dossier pour cet environnement, voulez-vous le créer ? (y/n)  ')
        if(choice == 'y'):
            # Création de l'environnement
            os.system('mkdir ' + env)
            os.chdir(env)
            os.system('mkdir weights')
            os.system('mkdir plots')
            wb = Workbook()
            wb.save('results.xlsx')
            os.chdir('..')
        elif(choice == 'n'):
            sys.exit()
    else:
        print('Environnement OK')


def check_xml(algo, env):
    filepath = '../' + algo + '/' + env + '/results.xlsx'
    print(filepath)
    wb = load_workbook(filepath)
    try:
        wb.save(filepath)
    except PermissionError as error:
        print(error)
        print('Le fichier excel est déjà ouvert ! Merci de le fermer avant le lancement d\'un script.')
        sys.exit()


def check_overwrite(algo, env, model):
    choice = 'null'
    weightpath = os.listdir('../' + algo + '/' + env + '/weights/')
    for weight in weightpath:
        if(weight.startswith(model)):
            while(choice != 'y' and choice != 'n'):
                choice = input('Un model porte déjà ce nom, voulez-vous écrire par dessus ? (y/n)  ')
            if(choice == 'y'):
                pass
            elif(choice == 'n'):
                sys.exit()